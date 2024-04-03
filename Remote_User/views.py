from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
from django.db.models import Q

# Create your views here.
from Remote_User.models import ClientRegister_Model,Clientreadings_Model,studentdata_Model


def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:

            enter = ClientRegister_Model.objects.get(username=username, password=password)
            request.session["userid"] = enter.id
            return redirect('Add_Data_Sets')
        except:
            pass

    return render(request,'RUser/login.html')

def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:

        return render(request,'RUser/Register1.html')


def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})

def Add_Data_Sets(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_Data_Sets.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        studentdata_Model.objects.all().delete()
        for r in range(1, active_sheet.max_row + 1):
            studentdata_Model.objects.create(
            Regno=active_sheet.cell(r, 1).value,
            names=active_sheet.cell(r, 2).value,
            Gender=active_sheet.cell(r, 3).value,
            Age=active_sheet.cell(r, 4).value,
            Height=active_sheet.cell(r, 5).value,
            Weight=active_sheet.cell(r, 6).value,
            Physical_Fit=active_sheet.cell(r, 7).value,
            cardio_Fit=active_sheet.cell(r, 8).value,
            Aerobic_Fit=active_sheet.cell(r, 9).value,
            Stress=active_sheet.cell(r, 10).value,
            Mental_Health=active_sheet.cell(r, 11).value,
            Intelligent=active_sheet.cell(r, 12).value,
            Executive_fun=active_sheet.cell(r, 13).value,
            Eating=active_sheet.cell(r, 14).value,
            Pyhsical_Activity=active_sheet.cell(r, 15).value,
            Sleep_pattern=active_sheet.cell(r, 16).value,
            Social_Tie=active_sheet.cell(r, 17).value,
            Time_Management=active_sheet.cell(r, 18).value,
            Stu_Class=active_sheet.cell(r, 19).value,
            Attendance=active_sheet.cell(r, 20).value,
            Library_Entry=active_sheet.cell(r, 21).value,
            Online_learning=active_sheet.cell(r, 22).value,
            Semester=active_sheet.cell(r, 23).value,
            ratings=0,
            usefulcounts=0,
            dislikes=0)

        return render(request, 'RUser/Add_Data_Sets.html', {"excel_data": excel_data})


def View_Student_Records(request):
    userid = request.session['userid']
    obj = studentdata_Model.objects.all()

    return render(request,'RUser/View_Student_Records.html',{'list_objects': obj})

def Search_Student_Data(request):
    value = request.POST.get('regno')
    obj1 = studentdata_Model.objects.all().filter(Q(Regno=value)| Q(names=value))

    return render(request, 'RUser/Search_Student_Data.html', {'list_objects': obj1})

